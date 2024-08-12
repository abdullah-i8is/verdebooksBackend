from distutils.log import debug
from flask import Flask, render_template, request, redirect, url_for, session, make_response
from flask_socketio import SocketIO, join_room, leave_room, emit
from flask_session import Session
from datetime import datetime
import mysql.connector as mysql
from flask import jsonify
from flask_cors import CORS
from datetime import datetime,timedelta
from datetime import timedelta
from datetime import date
import shutil
from gevent.pywsgi import WSGIServer
from openpyxl import Workbook
import openpyxl
import convertapi
from openpyxl.styles import Color, PatternFill, Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN
import pdfcrowd
from openpyxl.styles import Alignment
from random import randint, randrange
from werkzeug.datastructures import ImmutableMultiDict
from http.server import HTTPServer
from flask_cors import CORS
from werkzeug.utils import secure_filename
from io import BytesIO
import pdfplumber
import re
import random

app = Flask(__name__)
app.debug = True
app.config['SECRET_KEY'] = 'secret123123'
app.config['SESSION_TYPE'] = 'filesystem'



db_name = "Verdebooksdb"
db_password = "Hayat_3388171"
db_user = "hayat"
db_host = "ec2-15-222-247-146.ca-central-1.compute.amazonaws.com"


# db_name = "verdeinfiniti"
# db_password = ""
# db_user = "root"
# db_host = "localhost"


# db_name = "liveverde"
# db_password = ""
# db_user = "root"
# db_host = "verdebooks.com"

CORS(app)

def database():
    db = mysql.connect(host=db_host, user=db_user, passwd=db_password, database=db_name)
    cursor=db.cursor(buffered=True)
    return db,cursor
# RUN PAYROLL

def calculate(values):
    print(values)
    record={}
    RegCurrent=values["payRate"]*values["regHours"]
    record["RegCurrent"]=RegCurrent
    YTD=RegCurrent+values["YTDPrev"]
    record["YTD"]=YTD
    OTHours=values["OTHours"]
    record["OTHours"]=OTHours
    OTRate=values["payRate"]*1.5
    record["OTRate"]=OTRate
    OTCurrent=values["OTHours"]*OTRate
    record["OTCurrent"]=OTCurrent
    OTYTD=OTCurrent+values["OTYTDPrev"]
    record["OTYTD"]=OTYTD
    VACCurrent=(RegCurrent+OTCurrent+(values["stat"]*values["payRate"]))*0.04
    #VACCurrent=(values["payRate"]*values["stat"])*0.04
    record["VACCurrent"]=VACCurrent
    VACYTD=VACCurrent+values["VACYTDPrev"]
    record["VACYTD"]=VACYTD
    StatHours=values["stat"]
    record["StatHours"]=StatHours
    StatRate=values["statRate"]
    record["StatRate"]=StatRate
    StatCurrent=values["stat"]*values["payRate"]
    record["StatCurrent"]=StatCurrent
    StatYTD=StatCurrent+values["StatYTDPrev"]
    record["StatYTD"]=StatYTD
    shiftPremiumCurrent=values["shiftPremiumPayRate"]*values["shiftPremiumHours"]
    record["shiftPremiumCurrent"]=shiftPremiumCurrent
    shiftPremiumYTD=shiftPremiumCurrent+values["shiftPremiumYTD"]
    record["shiftPremiumYTD"]=shiftPremiumYTD
    print("Shift Premium C")
    print(shiftPremiumCurrent)
    if values["EIYTDPrev"]>=7454:
        IncomeTax=0.00
        record["IncomeTax"]=IncomeTax
        IncomeTaxYTD=values["IncomeTaxYTDPrev"]
        record["IncomeTaxYTD"]=IncomeTaxYTD
    else:
        IncomeTax=(RegCurrent+OTCurrent+VACCurrent+StatCurrent+shiftPremiumCurrent)*0.0505
        record["IncomeTax"]=IncomeTax
        IncomeTaxYTD=IncomeTax+values["IncomeTaxYTDPrev"]
        record["IncomeTaxYTD"]=IncomeTaxYTD
    if values["EIYTDPrev"]>=60300:
        EIYTD=values["EIYTDPrev"]
        record["EIYTD"]=EIYTD
        EI=0.00
        record["EI"]=EI
    else:
        EI=(RegCurrent+OTCurrent+VACCurrent+StatCurrent+shiftPremiumCurrent)*0.0158
        record["EI"]=EI
        EIYTD=EI+values["EIYTDPrev"]
        record["EIYTD"]=EIYTD
    if values["CPPYTDPrev"]>=3500:
        CPPYTD=values["CPPYTDPrev"]
        record["CPPYTD"]=CPPYTD
        CPP=0.00
        record["CPP"]=CPP
    else:
        CPP=(RegCurrent+OTCurrent+VACCurrent+StatCurrent+shiftPremiumCurrent)*0.0545
        record["CPP"]=CPP
        CPPYTD=CPP+values["CPPYTDPrev"]
        record["CPPYTD"]=CPPYTD
    TotalPayCurrent=RegCurrent+OTCurrent+VACCurrent+StatCurrent+shiftPremiumCurrent
    record["TotalPayCurrent"]=TotalPayCurrent
    TotalPayYTD=TotalPayCurrent+values["TotalPayYTDPrev"]
    record["TotalPayYTD"]=TotalPayYTD
    TotalTaxCurrent=IncomeTax+EI+CPP
    record["TotalTaxCurrent"]=TotalTaxCurrent
    TotalTaxYTD=TotalTaxCurrent+values["TotalTaxYTDPrev"]
    record["TotalTaxYTD"]=TotalTaxYTD
    NetPay=TotalPayCurrent-TotalTaxCurrent
    record["NetPay"]=NetPay
    return record

def transactionsManaager(payDate,payDate2,companyId):
    db, cursor = database()
    print("printingPayDate")
    print(payDate)

    query = "select id, date, balance from transactionalhistory where date = '" + str(
        payDate2) + "' and company_id = '" + str(companyId) + "' order by id desc limit 1"
    cursor.execute(query)
    result = cursor.fetchall()

    query3 = "select id, date, balance from transactionalhistory where date < '" + str(
        payDate2) + "' and company_id = '" + str(companyId) + "' order by date desc limit 1"
    cursor.execute(query3)
    result3 = cursor.fetchall()

    print("Printing Result")
    print(result)

    if result:
        query2 = "select id, date, balance from transactionalhistory where date = '" + str(
            result[0][1]) + "' and company_id = '" + str(companyId) + "' order by id desc limit 1"
        cursor.execute(query2)
        result2 = cursor.fetchall()
        print("Printing Result 2")
        print(result2)
        if result2:
            return result2[0]
        else:
            return result[0]
    elif result3:
        query2 = "select id, date, balance from transactionalhistory where date = '" + str(
            result3[0][1]) + "' and company_id = '" + str(companyId) + "' order by id desc limit 1"
        cursor.execute(query2)
        result2 = cursor.fetchall()
        print("Printing Result 2")
        print(result2)
        if result2:
            return result2[0]
        else:
            return result3[0]
    else:
        payDate = datetime.strptime(payDate, '%Y-%m-%d')
        depositeDate = payDate - timedelta(days=2)
        return (0, depositeDate, 0)

def updateBalance(payDate,lastBalance,withdraw,defaultDeposite):
    db,cursor=database()
    query="select id,withdrawal,balance,date from transactionalhistory where date > '"+str(payDate)+"'"
    cursor.execute(query)
    result=cursor.fetchall()
    ld=lastBalance
    if result:
        deposite=float(defaultDeposite)
        for row in result:
            balance=float(row[2])-float(withdraw)+float(deposite)
            if balance<0:
                rndBlnce=random.randint(5000,20000)
                description="DEPOSIT HAMILTON nextON 93310 490 next47943523 MB-DEP"
                query2="insert into transactionalhistory(date,description,withdrawal,deposite,balance) values(%s,%s,%s,%s,%s)"
                #depositeDate = datetime.strptime(str(row[3]),'%Y-%m-%d')
                depositeDate = datetime.strptime(str(row[3]).split()[0], '%Y-%m-%d')
                depositeDate = depositeDate - timedelta(days=2)
                cursor.execute(query2,(str(depositeDate),description,0.0,str(rndBlnce),str(ld+float(rndBlnce))))
                db.commit()
                query3="update transactionalhistory set balance=%s where id=%s"
                cursor.execute(query3,(str(float(row[2])+float(rndBlnce)-float(withdraw)),row[0]))
                db.commit()
                deposite=deposite+rndBlnce
                ld=float(row[2])-float(withdraw)+float(deposite)
            else:
                query3="update transactionalhistory set balance=%s where id=%s"
                cursor.execute(query3,(str(float(row[2])-float(withdraw)+float(deposite)),row[0]))
                db.commit()
                ld=float(row[2])-float(withdraw)+float(deposite)
    return "done"

def lastTransactionBalane():
    db,cursor=database()
    query="select balance from transactionalhistory order by id desc limit 1"
    cursor.execute(query)
    result=cursor.fetchone()
    print("Last Balance")
    print(result)
    return result[0]

@app.route("/api/runPayRoll",methods=["GET","POST"])
def runPayRoll():
    db,cursor=database()
    data=request.form
    print(data)
    employees=data["employees"]
    employees=employees.split(",")
    print("employees",employees)
    regularPayHours=data["payHours"]
    regularPayHours=regularPayHours.split(",")
    print('regularPayHours',regularPayHours)
    otHours=data["otHours"]
    otHours=otHours.split(",")
    print('otHours',otHours)
    stat=data["stat"]
    stat=stat.split(",")
    print('stat',stat)
    memo=data["memo"]
    memo=memo.split(",")
    print('memo',memo)
    shiftPremiumHours=data["statPremiumHours"]
    shiftPremiumHours=shiftPremiumHours.split(",")
    print("shiftPremiumHours",shiftPremiumHours)
    #payDate=data["payDate"]
    week=data["week"]
    week=week.split(" - ")
    print('week',week)
    payDate = datetime.strptime(week[1], "%Y-%m-%d") + timedelta(days=7)
    payDate=str(payDate)
    payDate=payDate[0:10]
    status={}
    generateDate=str(datetime.today().strftime('%Y-%m-%d'))
    tm=transactionsManaager(week[1],payDate,data["companyid"])
    balance=tm[2]
    print("Printing Employees")
    print(employees)
    for i in range(len(employees)):
        query="select id from stubs where employeeId=%s and company_id=%s and weekStart=%s and weekEnd=%s"
        values=(employees[i], data["companyid"], week[0],week[1])
        cursor.execute(query,values)
        result=cursor.fetchall()
        if result:status[str(employees[i])]="Already Generated"
        else:
            tempYear=week[0].split("-")
            year=tempYear[0]
            query1="select YTD,OTHours,OTYTD,VACYTD,StatYTD,EIYTD,CPPYTD,IncomeTaxYTD,TotalTaxYTD,TotalPayYTD,StatHours,StatRate,statPremiumYTD from stubs where employeeId=%s and company_id=%s and payDate < %s and weekStart like '%"+str(year)+"%' order by id desc limit 1"
            cursor.execute(query1, (employees[i], data["companyid"], payDate))
            result1=cursor.fetchone()
            print("result1",result1)
            if result:pass
            else:
                query1="select YTD,OTHours,OTYTD,VACYTD,StatYTD,EIYTD,CPPYTD,IncomeTaxYTD,TotalTaxYTD,TotalPayYTD,StatHours,StatRate,statPremiumYTD from stubs where employeeId=%s and company_id=%s and payDate < %s and weekStart like '%"+str(year)+"%' or NetPay=0 order by id desc limit 1"
                cursor.execute(query1,(employees[i], data["companyid"], payDate))
                result1=cursor.fetchone()
                print("result1",result1)
            if result1:
                query2="select name,payRate,shiftPremiumPayRate from employee where id=%s"
                cursor.execute(query2,(employees[i],))
                result2=cursor.fetchone()
                print(result2)
                if result2:
                    print("generating")
                    temp={"payRate":float(result2[1]),"regHours":float(regularPayHours[i]),"YTDPrev":float(result1[0]),"OTHours":float(otHours[i]),"OTYTDPrev":float(result1[2]),
                    "stat":float(stat[i]),"VACYTDPrev":float(result1[3]),"payDate":payDate,"StatYTDPrev":float(result1[4]),"IncomeTaxYTDPrev":float(result1[7]),"EIYTDPrev":float(result1[5]),
                    "CPPYTDPrev":float(result1[6]),"TotalTaxYTDPrev":float(result1[8]),"TotalPayYTDPrev":float(result1[9]),"statRate":float(result1[10]),"shiftPremiumPayRate":float(result2[2]),"shiftPremiumHours":float(shiftPremiumHours[i]),"shiftPremiumYTD":float(result1[12])}
                    record=calculate(temp)
                    record["name"]=result2[0]
                    record["employeeId"]=employees[0]
                    query3="insert into stubs(`company_id`, `generateDate`, `payDate`, `name`, `RegCurrent`, `YTD`, `OTHours`, `OTRate`, `OTCurrent`, `OTYTD`, `VACCurrent`, `VACYTD`, `StatHours`, `StatRate`, `StatYTD`, `IncomeTax`, `IncomeTaxYTD`, `EI`, `EIYTD`, `CPP`, `CPPYTD`, `TotalPayCurrent`, `TotalPayYTD`, `TotalTaxCurrent`, `TotalTaxYTD`, `NetPay`, `employeeId`, `status`,`weekStart`,`weekEnd`,`regHours`,`statPremiumHours`,`statPremiumCurrent`,`statPremiumYTD`,`memo`)  value(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                    values3=(data["companyid"], generateDate, payDate, result2[0], record["RegCurrent"], record["YTD"], record["OTHours"], record["OTRate"], record["OTCurrent"], record["OTYTD"], record["VACCurrent"],
                    record["VACYTD"], record["StatHours"], record["StatRate"], record["StatYTD"], record["IncomeTax"], record["IncomeTaxYTD"], record["EI"], record["EIYTD"], record["CPP"],
                    record["CPPYTD"], record["TotalPayCurrent"], record["TotalPayYTD"], record["TotalTaxCurrent"],record["TotalTaxYTD"], record["NetPay"], employees[i], "original",week[0],week[1],regularPayHours[i],shiftPremiumHours[i],record["shiftPremiumCurrent"],record["shiftPremiumYTD"],memo[i])
                    cursor.execute(query3,values3)
                    db.commit()
                    status[str(employees[i])]="Newly Generated"
                    currentBalance=float(balance)-float(record["NetPay"])
                    print(balance)
                    print(record["TotalPayCurrent"])
                    print("Printing Current Balance")
                    print(currentBalance)
                    count=0
                    defaultDeposite=0
                    rndBlnce=random.randint(5000,20000)
                    if currentBalance < 0:
                        while True:
                            currentBalance=float(currentBalance)+float(rndBlnce)
                            count=count+1
                            if currentBalance < 0:pass
                            else:break
                        description="DEPOSIT HAMILTON nextON 93310 490 next47943523 MB-DEP"
                        query8="insert into transactionalhistory(date,description,withdrawal,deposite,balance,company_id) values(%s,%s,%s,%s,%s,%s)"
                        payDate=datetime.strptime(str(payDate),'%Y-%m-%d')
                        depositeDate = payDate - timedelta(days=2)
                        cursor.execute(query8,(str(depositeDate),description,0,str(rndBlnce*count),str(float(balance)+(rndBlnce*count)),data["companyid"]))
                        db.commit()
                        defaultDeposite=rndBlnce*count
                    else:pass
                    query101="select id from stubs order by id desc limit 1"
                    cursor.execute(query101)
                    result101=cursor.fetchone()
                    #descriptions=["MISC PAYMENT nextTRANSFERWISE nextCANADA_____562D","ACCOUNTS PAYABLE nextTRAN FEE nextINTUIT CANADA U","ACCOUNTS PAYABLE next000001016983283 nextMAVERICK MACHINE nextAND HYDRAULI","ACCOUNTS PAYABLE next000001016983283 MAVERICK nextMACHINE AND nextHYDRAULI 1,178.36 32,648.93","ACCOUNTS PAYABLE nextDEPOSIT INTUIT nextCANADA U","BALANCE FORWARD","SERVICE CHARGE nextSCOTIA DIRECT nextPAYMENT","SD SETTLEMENT nextSD# 10460 nextFCN 0098 nextFCD 220303","BILL PAYMENT nextTXNFEE 1*$2 nextGOVERNMENT TAX nextPAYMENTS","ACCOUNTS PAYABLE nextTRAILERMASTER F","SD SETTLEMENT nextSD# 10460 FCN next0099 FCD 220304","ACCOUNTS PAYABLE nextTRAN FEE INTUIT nextCANADA U","SERVICE CHARGE","PAYROLL DEP. nextPYR WELD O nextCANADA INC","PAYROLL DEP. nextPYR WELD O nextCANADA INC"]
                    #selectedDes=random.randint(0,len(descriptions)-1)
                    query5="insert into transactionalhistory(date,description,withdrawal,deposite,balance,stubId, company_id) values(%s,%s,%s,%s,%s,%s,%s)"
                    rndNumber=random.randint(10000000,99999999)
                    values5=(str(payDate),"DEBIT MEMO next"+str(rndNumber)+" nextINTERAC E-TRANSFER",str(record["NetPay"]),0,str(currentBalance),result101[0], data["companyid"])
                    print(values5,"values5")
                    cursor.execute(query5,values5)
                   
                    db.commit()
                    updateBalance(payDate,currentBalance,record["NetPay"],defaultDeposite)
                    balance=lastTransactionBalane()
    response=[]
    print(status)
    print(employees)
    for i in range(len(employees)):
        query="select * from stubs where employeeId=%s and weekStart=%s and weekEnd=%s"
        values=(employees[i],week[0],week[1])
        cursor.execute(query,values)
        result=cursor.fetchone()
        if result:
            columns=("id","company_id", "generateDate", "payDate", "name", "RegCurrent", "YTD", "OTHours", "OTRate", "OTCurrent", "OTYTD", "VACCurrent", "VACYTD", "StatHours", "StatRate", "StatYTD", "IncomeTax", "IncomeTaxYTD", "EI", "EIYTD", "CPP", "CPPYTD", "TotalPayCurrent", "TotalPayYTD", "TotalTaxCurrent", "TotalTaxYTD", "NetPay", "employeeId", "status", "weekStart", "weekEnd", "regHours")
            temp={}
            for j in range(len(columns)):
                if j>0:
                    if columns[j]=="regHours":
                        try:temp[columns[j]]="{:.2f}".format(float(result[j])+float(result[6])+float(result[12]))
                        except:temp[columns[j]]=float(result[j])+float(result[6])+float(result[12])
                    else:
                        try:temp[columns[j]]="{:.2f}".format(float(result[j]))
                        except:temp[columns[j]]=result[j]
                else:temp[columns[j]]=result[j]
            temp["status"]=status[str(employees[i])]
            response.append(temp)
    return jsonify({"response":response})

#  FETCH EMPLOYEES

@app.route("/api/allEmployees",methods=['GET','POST'])
def allEmployees():
    db,cursor=database()
    data = request.form
    query="select id,paymentMethod,name,payRate,status,jobTitle,shiftPremiumStatus,shiftPremiumPayRate from employee where company_id=%s order by id desc"
    cursor.execute(query, (data["companyid"],))
    result=cursor.fetchall()
    response=[]
    for row in result:
        status=""
        if row[4]=="null" or row[4]==None:status="No Status"
        else:status=row[4]
        temp={"id":row[0],"paymentMethod":row[1],"name":row[2],"payRate":row[3],"status":status,"jobTitle":row[5],"shiftPremiumStatus":row[6],"shiftPremiumPayRate":row[7]}
        response.append(temp)
    return jsonify({"response":response})

# ADD EMPLOYEE

@app.route("/api/addOrganization", methods=["GET", "POST"])
def addOrganization():
    db,cursor=database()
    data=request.form
    query0="select * from organization where org_name=%s"
    cursor.execute(query0,(data["orgname"],))
    result0=cursor.fetchall()
    
    print(result0)
    if result0:return jsonify({"response":"exist"})
    else:
        query="INSERT INTO `organization`(`org_name`, `postal_code`, `street_address`, `city_Address`)  VALUES(%s,%s,%s,%s)"
        values=( data["orgname"], data["postalcode"], data["street"], data["city"])
        cursor.execute(query,values)
        db.commit()
        query2="Insert into `usersdata`(`email`, `password`, `status`, `organization`) values (%s,%s,%s,%s)"    
        values2=(data['email'], data['password'], "verified", data['orgname'])
        cursor.execute(query2, values2)
        db.commit()
# Assuming the operation was successful
        response_data = {
            "success": True,
            "message": "Account created successfully!",
            "data": None  # You can include any additional data here
        }
        return jsonify(response_data), 200


@app.route("/api/addEmployee",methods=["GET","POST"])
def addEmployee():
    db,cursor=database()
    data=request.form
    query0="select id from employee where email=%s and payRate=%s"
    cursor.execute(query0,(data['name'],data['payRate']))
    result0=cursor.fetchall()
    print("email")
    print(data["email"])
    print("result")
    print(result0)
    if result0:return jsonify({"response":"exist"})
    else:
        query="INSERT INTO `employee`(`company_id`,`name`, `jobTitle`, `status`, `hireDate`, `dob`, `workingLocation`, `accountHolder`,`bankName`, `accountNumber`, `branchName`, `bankLocation`, `address`, `town`, `postalCode`, `phn`, `phn2`, `gender`, `notes`, `mi`, `payRate`, `payType`, `vacPolicy`, `deduction`, `paymentMethod`, `email`,`shiftPremiumStatus`,`shiftPremiumPayRate`)  VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
        values=(data["companyid"], data["name"], data["jobTitle"], data["status"], data["hireDate"], data["dob"], data["workingLocation"], data["accountHolder"],data["bankName"],data["accountNumber"], data["branchName"], data["bankLocation"], data["address"], data["town"], data["postalCode"], data["phn"], data["phn2"],data["gender"], data["notes"], data["mi"], data["payRate"], data["payType"], data["vacPolicy"], data["deduction"], data["paymentMethod"], data["email"],data["shiftPremiumStatus"],data["shiftPremiumPayRate"])
        cursor.execute(query,values)
        db.commit()
        query3="select id from employee order by id desc limit 1"
        cursor.execute(query3)
        result3=cursor.fetchone()
        query2="insert into stubs(`company_id`,`generateDate`, `payDate`, `name`, `RegCurrent`, `YTD`, `OTHours`, `OTRate`, `OTCurrent`, `OTYTD`, `VACCurrent`, `VACYTD`, `StatHours`, `StatRate`, `StatYTD`, `IncomeTax`, `IncomeTaxYTD`, `EI`, `EIYTD`, `CPP`, `CPPYTD`, `TotalPayCurrent`, `TotalPayYTD`, `TotalTaxCurrent`, `TotalTaxYTD`, `NetPay`, `employeeId`, `status`,`weekStart`,`weekEnd`,`regHours`,`statPremiumHours`,`memo`)  value(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
        generateDate=str(datetime.today().strftime('%Y-%m-%d'))
        values2=(data["companyid"],generateDate,0,data["name"],0,0,0,0,0,0,0,0,0,1.5,0,0,0,0,0,0,0,0,0,0,0,0,result3[0],'initial',0,0,0,0,'')
        cursor.execute(query2,values2)
        db.commit()
        return jsonify({"response":"done"})

# FETCH CHEQUE

@app.route("/api/employeeStubList/<data>", methods=["GET","POST"])
def employeeChequeList(data):
    db,cursor=database()
    query="select payDate,name,NetPay,id,totalPayCurrent from stubs where employeeId=%s order by id desc"
    value=(data,)
    cursor.execute(query,value)
    result=cursor.fetchall()
    response=[]
    for row in result:
        if row[0]==0 or row[0]=="0":pass
        else:
            tempDate=str(row[0])
            temp={"payDate":tempDate[0:10],"name":row[1],"totalPay":"{:.2f}".format(float(row[4])),"netPay":"{:.2f}".format(float(row[2])),"paymentMethod":"bank","id":row[3]}
            response.append(temp)
    return jsonify({"response":response})

# FETCH EMPLOYEE DETAIL

@app.route("/api/employeeProfile/<data>",methods=["GET","POST"])
def employeeProfile(data):
    db,cursor=database()
    query="select * from employee where id=%s"
    cursor.execute(query,(data,))
    result=cursor.fetchone()
    print(result)
    if result:
        response={"name":result[2],"jobTitle":result[3],"status":result[4],"hireDate":result[5],"dob":result[6],"workingLocation":result[7],"accountHolder":result[8],"bankName":result[9],"accountNumber":result[10],"branchName":result[11]
        ,"bankLocation":result[12],"address":result[13],"town":result[14],"postalCode":result[15],"phn":result[16],"phn2":result[17],"gender":result[18],"notes":result[19],"mi":result[20],"payRate":result[21],"payType":result[22],
        "vacPolicy":result[23],"deduction":result[24],"paymentMethod":result[25],"email":result[26],"shiftPremiumStatus":result[27],"shiftPremiumPayRate":result[28]}
        print(response,"resultemployeeresponse")

        return jsonify({"response":response})
    else:
        return jsonify({"response":""})

# FETCH DATES

def numOfDays(date1, date2):
    return (date2-date1).days

@app.route("/api/getDates/<year>",methods=["GET","POST"])
def getDates(year):
    # Driver program
    start_date = date(2000,1,7)
    date2 = date.today()
    print(date2.year)
    days=numOfDays(start_date, date2)
    print(days)
    days=int(days/7)
    print(days)
    start_date=str(start_date)
    start_date=datetime.strptime(start_date, "%Y-%m-%d").strftime("%Y-%m-%d")
    fridays=[]
    dateRanges=[]
    count=0
    for i in range(days):
        previous=str(datetime.strptime(start_date, "%Y-%m-%d")+timedelta(days=1))
        friday=datetime.strptime(start_date, "%Y-%m-%d")+timedelta(days=7)
        generatedDate=str(friday)
        generatedDate=generatedDate[0:10]
        friday=datetime.strptime(str(friday.date()), "%Y-%m-%d").strftime("%Y-%m-%d")
        start_date=friday
        dateRange=previous[0:10] +" - " + str(generatedDate)
        if dateRange=="":
            pass
        else:
            if year in str(generatedDate):
                fridays.append(str(generatedDate))
                dateRanges.append(dateRange)
                count=count+1
    print(fridays)
    dateRanges.reverse()
    print(dateRanges)
    return jsonify({"dateRanges":dateRanges,"count":count})

# LOGIN


@app.route("/api/login",methods=["GET","POST"])
def login():
    db,cursor=database()
    data=request.form
    query="select id,password,organization from usersdata where email=%s"
    cursor.execute(query,(data["email"],))
    result=cursor.fetchone()
    response=""
    if result:
        if result[1]==data["password"]:
            response={"response":"Success","company":result[2], "id":result[0]}
        else:
            response={"response":"Invalid Password"}
    else:
        response={"response":"Invalid Email or Password"}
    return jsonify(response)




# PRINT STUB
@app.route("/api/printStub/<data>",methods=["GET","POST"])
def printStub(data):
    db,cursor=database()
    userid=data.split('-')[0]
    organization=data.split('-')[1]
    query3="select name,weekEnd,weekStart,payDate,NetPay,regHours,RegCurrent,YTD,VACCurrent,VACYTD,StatHours,StatRate,StatYTD,IncomeTax,IncomeTaxYTD,EI,EIYTD,TotalPayCurrent,TotalPayYTD,TotalTaxCurrent,TotalTaxYTD,OTHours,employeeId,CPP,CPPYTD,OTHours,OTYTD,statPremiumHours,statPremiumCurrent,statPremiumYTD, memo from stubs where id=%s"
    cursor.execute(query3,(userid,))
    record=cursor.fetchone()
    if record:
        query2="select address,town,payRate,shiftPremiumPayRate,postalCode, town from employee where id=%s"
        cursor.execute(query2,(record[22],))
        result2=cursor.fetchone()
        query4 = "select * from organization where org_name=%s"
        cursor.execute(query4, (organization,))
        recordorg = cursor.fetchone()

        postal_code = result2[4] if result2[4] != "null" else ""
        stub={"payDateTitle":"PAY DATE: "+str(record[3]),
              "orgname":recordorg[1],
              "orgpostalcode":recordorg[2],
              "orgstreet":recordorg[3],
              "orgcity":recordorg[4],
	"netPayTitle":"NET PAY: $"+str("{:.2f}".format(float(record[4]))),
        "name":record[0],
        "address":result2[0],
        "town":result2[1],
        "postalCode":postal_code,
        "periodStart":record[2],
        "periodEnd":record[1],
        "payDate":str(record[3]),
	"totalHours":str(float(record[5])+float(record[21])+float(record[10])),
        "netPay":"$"+str("{:.2f}".format(float(record[4]))),
        "regularPayCurrentHours":str("{:.2f}".format(float(record[5]))),
        "currentRegularPayRate":str("{:.2f}".format(float(result2[2]))),
        "currentRegularPay":str("{:.2f}".format(float(record[5])*float(result2[2]))),
        "currentRegularPayYTD":str("{:.2f}".format(float(record[7]))),
        "vacCurrentPay":str("{:.2f}".format(float(record[8]))),
        "vacPayYTD":str("{:.2f}".format(float(record[9]))),
        "statCurrentHours":str("{:.2f}".format(float(record[10]))),
        "statCurrentRate":str("{:.2f}".format(float(result2[2]))),
	"statCurrentPay":str("{:.2f}".format(float(record[10])*float(result2[2]))),
	"statCurrentYTD":str("{:.2f}".format(float(record[12]))),
	"incomeTax":str("{:.2f}".format(float(record[13]))),
        "incomeTaxYTD":str("{:.2f}".format(float(record[14]))),
        "ei":str("{:.2f}".format(float(record[15]))),
	"eiYTD":str("{:.2f}".format(float(record[16]))),
        "totalPayCurrent":str("{:.2f}".format(float(record[17]))),
	"totalPayYTD":str("{:.2f}".format(float(record[18]))),
	"totalTaxCurrent":str("{:.2f}".format(float(record[19]))),
	"totalTaxYTD":str("{:.2f}".format(float(record[20]))),
        "CPP":str("{:.2f}".format(float(record[23]))),
        "CPPYTD":str("{:.2f}".format(float(record[24]))),
        "OTHours":str("{:.2f}".format(float(record[25]))),
        "OTCurrentRate":str("{:.2f}".format(float(result2[2])*1.5)),
        "OTCurrentPay":str("{:.2f}".format(float(record[25])*(float(result2[2])*1.5))),
        "OTYTD":str("{:.2f}".format(float(record[26]))),
        "SPHours":str("{:.2f}".format(float(record[27]))),
        "SPCurrentPay":str("{:.2f}".format(float(record[28]))),
        "SPYTD":str("{:.2f}".format(float(record[29]))),
        "SPCurrentRate":str("{:.2f}".format(float(result2[3]))),
        "Memo":record[30]
	}
        return render_template("generateSlip.html",stub=stub)
    else:
        return jsonify({"response":"Invalid Id"})

#PRINT MULTIPLE STUBS

@app.route("/api/getMultipleStubs/<data>",methods=["GET","POST"])
def getMultpleStubs(data):
    data=data.split(",")
    data.sort()
    return render_template("stubs.html",stubs=data)

@app.route("/api/printMultipleStub/<data>",methods=["GET","POST"])
def printMultipleStub(data):
    db,cursor=database()
    query3="select name,weekEnd,weekStart,payDate,NetPay,regHours,RegCurrent,YTD,VACCurrent,VACYTD,StatHours,StatRate,StatYTD,IncomeTax,IncomeTaxYTD,EI,EIYTD,TotalPayCurrent,TotalPayYTD,TotalTaxCurrent,TotalTaxYTD,OTHours,employeeId,CPP,CPPYTD,OTHours,OTYTD,statPremiumHours,statPremiumCurrent,statPremiumYTD from stubs where id=%s"
    cursor.execute(query3,(data,))
    record=cursor.fetchone()
    if record:
        query2="select address,town,payRate,shiftPremiumPayRate from employee where id=%s"
        cursor.execute(query2,(record[22],))
        result2=cursor.fetchone()
        stub={"payDateTitle":"PAY DATE: "+str(record[3]),
	"netPayTitle":"NET PAY: $"+str("{:.2f}".format(float(record[4]))),
        "name":record[0],
        "address":result2[0],
        "town":result2[1],
        "periodStart":record[2],
        "periodEnd":record[1],
        "payDate":str(record[3]),
	"totalHours":str(float(record[5])+float(record[21])+float(record[10])),
        "netPay":"$"+str("{:.2f}".format(float(record[4]))),
        "regularPayCurrentHours":str("{:.2f}".format(float(record[5]))),
        "currentRegularPayRate":str("{:.2f}".format(float(result2[2]))),
        "currentRegularPay":str("{:.2f}".format(float(record[5])*float(result2[2]))),
        "currentRegularPayYTD":str("{:.2f}".format(float(record[7]))),
        "vacCurrentPay":str("{:.2f}".format(float(record[8]))),
        "vacPayYTD":str("{:.2f}".format(float(record[9]))),
        "statCurrentHours":str("{:.2f}".format(float(record[10]))),
        "statCurrentRate":str("{:.2f}".format(float(result2[2]))),
	"statCurrentPay":str("{:.2f}".format(float(record[10])*float(result2[2]))),
	"statCurrentYTD":str("{:.2f}".format(float(record[12]))),
	"incomeTax":str("{:.2f}".format(float(record[13]))),
        "incomeTaxYTD":str("{:.2f}".format(float(record[14]))),
        "ei":str("{:.2f}".format(float(record[15]))),
	"eiYTD":str("{:.2f}".format(float(record[16]))),
        "totalPayCurrent":str("{:.2f}".format(float(record[17]))),
	"totalPayYTD":str("{:.2f}".format(float(record[18]))),
	"totalTaxCurrent":str("{:.2f}".format(float(record[19]))),
	"totalTaxYTD":str("{:.2f}".format(float(record[20]))),
        "CPP":str("{:.2f}".format(float(record[23]))),
        "CPPYTD":str("{:.2f}".format(float(record[24]))),
        "OTHours":str("{:.2f}".format(float(record[25]))),
        "OTCurrentRate":str("{:.2f}".format(float(result2[2])*1.5)),
        "OTCurrentPay":str("{:.2f}".format(float(record[25])*(float(result2[2])*1.5))),
        "OTYTD":str("{:.2f}".format(float(record[26]))),
        "SPHours":str("{:.2f}".format(float(record[27]))),
        "SPCurrentPay":str("{:.2f}".format(float(record[28]))),
        "SPYTD":str("{:.2f}".format(float(record[29]))),
        "SPCurrentRate":str("{:.2f}".format(float(result2[3])))
	}
        return render_template("generateMultipleSlips.html",stub=stub)
    else:
        return jsonify({"response":"Invalid Id"})



# TRANSACTIONAL HISTORY

@app.route("/api/transactionalhistory",methods=["GET","POST"])
def transactionalhistory():
    db,cursor=database()
    query="select id,date,description,deposite,balance from transactionalhistory"
    cursor.execute(query)
    result=cursor.fetchall()
    if result:
        response=[]
        for row in result:
            temp={"id":row[0],"date":str(row[1]),"description":row[2],"deposite":row[3],"balance":row[4]}
            response.append(temp)
        return jsonify({"response":response})
    else:
        return jsonify({"response":"No Data Found"})
# GENERATE STATEMENT
@app.route("/api/generateStatement/<month>/<year>",methods=["GET","POST"])
def generateStatement(month, year):
    tempMonth = int(month)
    year = int(year)
    if tempMonth == 1:
        last_month = 12
        last_year = year - 1
    else:
        last_month = tempMonth - 1
        last_year = year
    last_month_str = f"{last_year}-{last_month:02}"
    db, cursor = database()
    query = f"SELECT * FROM transactionalhistory WHERE date LIKE '{last_month_str}%' ORDER BY date DESC, id DESC limit 1"
    cursor.execute(query)
    prvresult = cursor.fetchall()
    if not prvresult:
        if last_month == 1:
            last_month = 12
            last_year = year - 1
        else:
            last_month -= 1
        last_month_str = f"{last_year}-{last_month:02}"
        query = f"SELECT * FROM transactionalhistory WHERE date LIKE '{last_month_str}%' ORDER BY date DESC, id DESC limit 1"
        cursor.execute(query)
        prvresult = cursor.fetchall()
    print("Last transaction data:", prvresult)
    if tempMonth<10:
        tempMonth="0"+str(tempMonth)
    month=int(month)
    print("tempMonth",tempMonth)
    shutil.copy("/var/www/verdebooks/main.xlsx", "/var/www/verdebooks/webapp/statements/"+str(month)+"-"+str(year)+".xlsx")
    db,cursor=database()
    query="select * from transactionalhistory where date LIKE '%"+str(year)+"-"+str(month)+"%' OR date LIKE '%"+str(year)+"-"+str(tempMonth)+"%' order by date,id"
    cursor.execute(query)
    result=cursor.fetchall()
    print("result",result)
    xfile = openpyxl.load_workbook("/var/www/verdebooks/webapp/statements/"+str(month)+"-"+str(year)+".xlsx")
    sheet = xfile.get_sheet_by_name('Sheet1')
    sheet.page_margins.left=0.25
    sheet.page_margins.right=0.1
    sheet.page_margins.top=0.25
    sheet.page_margins.bottom=0.25
    img = openpyxl.drawing.image.Image('logo.png')
    sheet.add_image(img,'A1')
    cur=23
    noOfDebit=0
    noOfCredit=0
    debitTotal=0
    creditTotal=0
    tempDebitTotal=0
    tempNoOfDebit=0
    tempCreditTotal=0
    tempNoOfCredit=0
    statementPage=1
    greyFill = PatternFill(start_color='ECECEC',end_color='ECECEC',fill_type='solid')
    count=0
    balance=0
    align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    mediumBorderTop = Border(top=Side(style=BORDER_THIN, color='000000'))
    mediumBorderBottom = Border(bottom=Side(style=BORDER_THIN, color='000000'))
    mediumBorderToBottom=Border(top=Side(style=BORDER_THIN, color='000000'),bottom=Side(style=BORDER_THIN, color='000000'))
    titleFont=Font(size=14)
    previousMonth=int(month)-1
    temp=year
    if month==1:
       previousMonth=12
       temp=int(year)-1
    sheet['H13'] = str(previousMonth)+"-"+str(temp)
    sheet['J13']=str(month)+"-"+str(year)
    transCounter=1
    emptyStatement=True
    totalStatementPages=(len(result)-4)/7
    print("totalStatementPages",totalStatementPages)
    if totalStatementPages<=0:totalStatementPages=1
    else:
        temp=round(totalStatementPages)
        if temp>=totalStatementPages:totalStatementPages=temp+2
        else:totalStatementPages=temp+2
    if prvresult:
    #  prvresult[0] = ("2023-11-03", "forward balance", 0.0, 0.0, 0.0)
     row = prvresult[0]
     # Apply formatting to the data
     sheet['A' + str(cur)].alignment = align
     sheet['B' + str(cur)].alignment = align
     sheet['C' + str(cur)].alignment = align
     sheet['F' + str(cur)].alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
     sheet['H' + str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
     sheet['J' + str(cur)].alignment = align
     sheet.column_dimensions['B'].width = 11
     sheet.row_dimensions[cur].height = 50
     sheet.column_dimensions['C'].width = 1
     ddd = str(row[1])
     sheet['A' + str(cur)] = ddd[0:10]
     fbd = "BALANCE FORWARD"
     temp = fbd.replace("next", "\n")
     try:
         sheet['B' + str(cur)] = temp
     except:
         sheet['B' + str(cur)] = temp

     sheet.merge_cells("B" + str(cur) + ":" + "D" + str(cur))
     sheet['C' + str(cur)].alignment = align
     with0 = 0.0
     try:
         sheet['F' + str(cur)] = "{:.2f}".format(with0)
     except:
         sheet['F' + str(cur)] = with0

     try:
         sheet['H' + str(cur)] = "{:.2f}".format(with0)
     except:
         sheet['H' + str(cur)] = with0

     try:
         sheet['J' + str(cur)] = "{:.2f}".format(float(row[5]))
     except:
         sheet['J' + str(cur)] = row[5]

     cur += 2
    for row in result:
        emptyStatement=False
        if cur==31:
            transCounter=1
            cur=cur+1
            #sheet.row_dimensions[cur].height=30
            sheet['A'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            sheet['A'+str(cur)].border = mediumBorderTop
            sheet['A'+str(cur)].font = Font(bold=True)
            sheet['A'+str(cur+1)].border = mediumBorderBottom
            sheet['G'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            sheet['G'+str(cur)].border = mediumBorderTop
            sheet['G'+str(cur)].font = Font(bold=True)
            sheet['G'+str(cur+1)].border = mediumBorderBottom
            sheet['C'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            sheet['C'+str(cur)].border = mediumBorderTop
            sheet['C'+str(cur+1)].border = mediumBorderBottom
            sheet['C'+str(cur)].font = Font(bold=True)
            sheet['I'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            sheet['I'+str(cur)].border = mediumBorderTop
            sheet['I'+str(cur)].font = Font(bold=True)
            sheet['I'+str(cur+1)].border = mediumBorderBottom
            sheet['A'+str(cur)]="No. of Debits"
            sheet['A'+str(cur+1)]=str(noOfDebit)
            sheet.merge_cells("A"+str(cur)+":"+"B"+str(cur))
            sheet.merge_cells("A"+str(cur+1)+":"+"B"+str(cur+1))
            sheet['G'+str(cur)]="No. of Credits"
            sheet['G'+str(cur+1)]=str(noOfCredit)
            sheet.merge_cells("G"+str(cur)+":"+"H"+str(cur))
            sheet.merge_cells("G"+str(cur+1)+":"+"H"+str(cur+1))
            sheet['C'+str(cur)]="Total Amount - Debits \n$"
            sheet['C'+str(cur+1)]=str("{:.2f}".format(float(debitTotal)))
            sheet.merge_cells("C"+str(cur)+":"+"F"+str(cur))
            sheet.merge_cells("C"+str(cur+1)+":"+"F"+str(cur+1))
            sheet['I'+str(cur)]="Total Amount - Credits \n$"
            sheet['I'+str(cur+1)]=str("{:.2f}".format(creditTotal))
            sheet.merge_cells("I"+str(cur)+":"+"J"+str(cur))
            sheet.merge_cells("I"+str(cur+1)+":"+"J"+str(cur+1))
            cur=cur+3
            sheet['J'+str(35*int(statementPage))]=str(statementPage)+" of "+str(totalStatementPages)
            statementPage=statementPage+1
            cur=cur+2
            for i in range(1,15):
                if i==1:
                    img = openpyxl.drawing.image.Image('logo.png')
                    sheet.add_image(img,'A'+str(cur))
                    cur=cur+3
                else:
                    sheet['A'+str(cur)]=sheet['A'+str(i)].value
                    sheet['A'+str(cur)].font=titleFont
                    sheet['B'+str(cur)]=sheet['B'+str(i)].value
                    sheet['B'+str(cur)].font=titleFont
                    sheet['C'+str(cur)]=sheet['C'+str(i)].value
                    sheet['C'+str(cur)].font=titleFont
                    sheet['D'+str(cur)]=sheet['D'+str(i)].value
                    sheet['D'+str(cur)].font=titleFont
                    sheet['E'+str(cur)]=sheet['E'+str(i)].value
                    sheet['E'+str(cur)].font=titleFont
                    sheet['F'+str(cur)]=sheet['F'+str(i)].value
                    sheet['F'+str(cur)].font=titleFont
                    sheet['G'+str(cur)]=sheet['G'+str(i)].value
                    sheet['G'+str(cur)].font=titleFont
                    sheet['H'+str(cur)]=sheet['H'+str(i)].value
                    sheet['H'+str(cur)].font=titleFont
                    sheet['I'+str(cur)]=sheet['I'+str(i)].value
                    sheet['I'+str(cur)].font=titleFont
                    sheet['J'+str(cur)]=sheet['J'+str(i)].value
                    sheet['J'+str(cur)].font=titleFont
                    if i==5:
                        sheet['A'+str(cur)].border = mediumBorderBottom
                        sheet['B'+str(cur)].border = mediumBorderBottom
                        sheet['C'+str(cur)].border = mediumBorderBottom
                        sheet['D'+str(cur)].border = mediumBorderBottom
                        sheet['E'+str(cur)].border = mediumBorderBottom
                        sheet['F'+str(cur)].border = mediumBorderBottom
                        sheet['G'+str(cur)].border = mediumBorderBottom
                        sheet['H'+str(cur)].border = mediumBorderBottom
                        sheet['I'+str(cur)].border = mediumBorderBottom
                        sheet['J'+str(cur)].border = mediumBorderBottom
                    elif i==12:
                        sheet['A'+str(cur)].font = Font(bold=True)
                        sheet['B'+str(cur)].font = Font(bold=True)
                        sheet['C'+str(cur)].font = Font(bold=True)
                        sheet['D'+str(cur)].font = Font(bold=True)
                        sheet['E'+str(cur)].font = Font(bold=True)
                        sheet['F'+str(cur)].font = Font(bold=True)
                        sheet['G'+str(cur)].font = Font(bold=True)
                        sheet['H'+str(cur)].font = Font(bold=True)
                        sheet['I'+str(cur)].font = Font(bold=True)
                        sheet['J'+str(cur)].font = Font(bold=True)
                    else:pass
                    cur=cur+1
            sheet['A'+str(cur)]=sheet['A21'].value
            sheet['A'+str(cur)].font=Font(bold=True)
            sheet['B'+str(cur)]=sheet['B21'].value
            sheet['B'+str(cur)].font=Font(bold=True)
            sheet['C'+str(cur)]=sheet['C21'].value
            sheet['C'+str(cur)].font=Font(bold=True)
            sheet['D'+str(cur)]=sheet['D21'].value
            sheet['D'+str(cur)].font=Font(bold=True)
            sheet['E'+str(cur)]=sheet['E21'].value
            sheet['E'+str(cur)].font=Font(bold=True)
            sheet['F'+str(cur)]=sheet['F21'].value
            sheet['F'+str(cur)].font=Font(bold=True)
            sheet['G'+str(cur)]=sheet['G21'].value
            sheet['G'+str(cur)].font=Font(bold=True)
            sheet['H'+str(cur)]=sheet['H21'].value
            sheet['H'+str(cur)].font=Font(bold=True)
            sheet['I'+str(cur)]=sheet['I21'].value
            sheet['I'+str(cur)].font=Font(bold=True)
            sheet['J'+str(cur)]=sheet['J21'].value
            sheet['J'+str(cur)].font=Font(bold=True)
            sheet['A'+str(cur)].border = mediumBorderToBottom
            sheet['B'+str(cur)].border = mediumBorderToBottom
            sheet['C'+str(cur)].border = mediumBorderToBottom
            sheet['D'+str(cur)].border = mediumBorderToBottom
            sheet['E'+str(cur)].border = mediumBorderToBottom
            sheet['F'+str(cur)].border = mediumBorderToBottom
            sheet['G'+str(cur)].border = mediumBorderToBottom
            sheet['H'+str(cur)].border = mediumBorderToBottom
            sheet['I'+str(cur)].border = mediumBorderToBottom
            sheet['J'+str(cur)].border = mediumBorderToBottom
            tempDebitTotal=0
            tempNoOfDebit=0
            tempCreditTotal=0
            tempNoOfCredit=0
            cur=cur+2
        else:pass
        sheet['A'+str(cur)].alignment = align
        sheet['B'+str(cur)].alignment = align
        sheet['C'+str(cur)].alignment = align
        sheet['F'+str(cur)].alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
        sheet['H'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        sheet['J'+str(cur)].alignment = align
        sheet.column_dimensions['B'].width=11
        sheet.row_dimensions[cur].height=50
        sheet.column_dimensions['C'].width=1
        ddd=str(row[1])
        sheet['A'+str(cur)] = ddd[0:10]
        temp = row[2].replace("next","\n")
        try:sheet['B'+str(cur)] = temp
        except:sheet['B'+str(cur)] = temp
        sheet.merge_cells("B"+str(cur)+":"+"D"+str(cur))
        sheet['C'+str(cur)].alignment = align
        try:sheet['F'+str(cur)] = "{:.2f}".format(float(row[3]))
        except:sheet['F'+str(cur)] = row[3]
        try:sheet['H'+str(cur)] = "{:.2f}".format(float(row[4]))
        except:sheet['H'+str(cur)] = row[4]
        try:sheet['J'+str(cur)] = "{:.2f}".format(float(row[5]))
        except:sheet['J'+str(cur)] = row[5]
        if row[3]!="":
            debitTotal=debitTotal+float(row[3])
            noOfDebit=noOfDebit+1
            tempDebitTotal=tempDebitTotal+float(row[3])
            tempNoOfDebit=tempNoOfDebit+1
        if row[4]!="":
            creditTotal=creditTotal+float(row[4])
            noOfCredit=noOfCredit+1
            tempCreditTotal=tempCreditTotal+float(row[4])
            tempNoOfCredit=tempNoOfCredit+1
        if count%2==0:
            sheet['A'+str(cur)].fill = greyFill
            sheet['B'+str(cur)].fill = greyFill
            sheet['C'+str(cur)].fill = greyFill
            sheet['D'+str(cur)].fill = greyFill
            sheet['E'+str(cur)].fill = greyFill
            sheet['F'+str(cur)].fill = greyFill
            sheet['G'+str(cur)].fill = greyFill
            sheet['H'+str(cur)].fill = greyFill
            sheet['I'+str(cur)].fill = greyFill
            sheet['J'+str(cur)].fill = greyFill
        try:balance = "{:.2f}".format(float(row[5]))
        except:balance = row[5]
        if transCounter%6==0:
            cur=cur+2
            #sheet.row_dimensions[cur].height=30
            sheet['A'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            sheet['A'+str(cur)].border = mediumBorderTop
            sheet['A'+str(cur)].font = Font(bold=True)
            sheet['A'+str(cur+1)].border = mediumBorderBottom
            sheet['G'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            sheet['G'+str(cur)].border = mediumBorderTop
            sheet['G'+str(cur)].font = Font(bold=True)
            sheet['G'+str(cur+1)].border = mediumBorderBottom
            sheet['C'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            sheet['C'+str(cur)].border = mediumBorderTop
            sheet['C'+str(cur+1)].border = mediumBorderBottom
            sheet['C'+str(cur)].font = Font(bold=True)
            sheet['I'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            sheet['I'+str(cur)].border = mediumBorderTop
            sheet['I'+str(cur)].font = Font(bold=True)
            sheet['I'+str(cur+1)].border = mediumBorderBottom
            sheet['A'+str(cur)]="No. of Debits"
            sheet['A'+str(cur+1)]=str(tempNoOfDebit)
            sheet.merge_cells("A"+str(cur)+":"+"B"+str(cur))
            sheet.merge_cells("A"+str(cur+1)+":"+"B"+str(cur+1))
            sheet['G'+str(cur)]="No. of Credits"
            sheet['G'+str(cur+1)]=str(tempNoOfCredit)
            sheet.merge_cells("G"+str(cur)+":"+"H"+str(cur))
            sheet.merge_cells("G"+str(cur+1)+":"+"H"+str(cur+1))
            sheet['C'+str(cur)]="Total Amount - Debits \n$"
            sheet['C'+str(cur+1)]=str("{:.2f}".format(float(tempDebitTotal)))
            sheet.merge_cells("C"+str(cur)+":"+"F"+str(cur))
            sheet.merge_cells("C"+str(cur+1)+":"+"F"+str(cur+1))
            sheet['I'+str(cur)]="Total Amount - Credits \n$"
            sheet['I'+str(cur+1)]=str("{:.2f}".format(tempCreditTotal))
            sheet.merge_cells("I"+str(cur)+":"+"J"+str(cur))
            sheet.merge_cells("I"+str(cur+1)+":"+"J"+str(cur+1))
            cur=cur+3
            tempDebitTotal=0
            tempNoOfDebit=0
            tempCreditTotal=0
            tempNoOfCredit=0
            cur=cur+1
            try:sheet['J'+str(35*int(statementPage))]=str(statementPage)+" of "+str(totalStatementPages)
            except:pass
            statementPage=statementPage+1
            cur=cur+2
            for i in range(1,15):
                if i==1:
                    img = openpyxl.drawing.image.Image('logo.png')
                    sheet.add_image(img,'A'+str(cur))
                    cur=cur+3
                else:
                    sheet['A'+str(cur)]=sheet['A'+str(i)].value
                    sheet['A'+str(cur)].font=titleFont
                    sheet['B'+str(cur)]=sheet['B'+str(i)].value
                    sheet['B'+str(cur)].font=titleFont
                    sheet['C'+str(cur)]=sheet['C'+str(i)].value
                    sheet['C'+str(cur)].font=titleFont
                    sheet['D'+str(cur)]=sheet['D'+str(i)].value
                    sheet['D'+str(cur)].font=titleFont
                    sheet['E'+str(cur)]=sheet['E'+str(i)].value
                    sheet['E'+str(cur)].font=titleFont
                    sheet['F'+str(cur)]=sheet['F'+str(i)].value
                    sheet['F'+str(cur)].font=titleFont
                    sheet['G'+str(cur)]=sheet['G'+str(i)].value
                    sheet['G'+str(cur)].font=titleFont
                    sheet['H'+str(cur)]=sheet['H'+str(i)].value
                    sheet['H'+str(cur)].font=titleFont
                    sheet['I'+str(cur)]=sheet['I'+str(i)].value
                    sheet['I'+str(cur)].font=titleFont
                    sheet['J'+str(cur)]=sheet['J'+str(i)].value
                    sheet['J'+str(cur)].font=titleFont
                    if i==5:
                        sheet['A'+str(cur)].border = mediumBorderBottom
                        sheet['B'+str(cur)].border = mediumBorderBottom
                        sheet['C'+str(cur)].border = mediumBorderBottom
                        sheet['D'+str(cur)].border = mediumBorderBottom
                        sheet['E'+str(cur)].border = mediumBorderBottom
                        sheet['F'+str(cur)].border = mediumBorderBottom
                        sheet['G'+str(cur)].border = mediumBorderBottom
                        sheet['H'+str(cur)].border = mediumBorderBottom
                        sheet['I'+str(cur)].border = mediumBorderBottom
                        sheet['J'+str(cur)].border = mediumBorderBottom
                    elif i==12:
                        sheet['A'+str(cur)].font = Font(bold=True)
                        sheet['B'+str(cur)].font = Font(bold=True)
                        sheet['C'+str(cur)].font = Font(bold=True)
                        sheet['D'+str(cur)].font = Font(bold=True)
                        sheet['E'+str(cur)].font = Font(bold=True)
                        sheet['F'+str(cur)].font = Font(bold=True)
                        sheet['G'+str(cur)].font = Font(bold=True)
                        sheet['H'+str(cur)].font = Font(bold=True)
                        sheet['I'+str(cur)].font = Font(bold=True)
                        sheet['J'+str(cur)].font = Font(bold=True)
                    else:pass
                    cur=cur+1
            sheet['A'+str(cur)]=sheet['A21'].value
            sheet['A'+str(cur)].font=Font(bold=True)
            sheet['B'+str(cur)]=sheet['B21'].value
            sheet['B'+str(cur)].font=Font(bold=True)
            sheet['C'+str(cur)]=sheet['C21'].value
            sheet['C'+str(cur)].font=Font(bold=True)
            sheet['D'+str(cur)]=sheet['D21'].value
            sheet['D'+str(cur)].font=Font(bold=True)
            sheet['E'+str(cur)]=sheet['E21'].value
            sheet['E'+str(cur)].font=Font(bold=True)
            sheet['F'+str(cur)]=sheet['F21'].value
            sheet['F'+str(cur)].font=Font(bold=True)
            sheet['G'+str(cur)]=sheet['G21'].value
            sheet['G'+str(cur)].font=Font(bold=True)
            sheet['H'+str(cur)]=sheet['H21'].value
            sheet['H'+str(cur)].font=Font(bold=True)
            sheet['I'+str(cur)]=sheet['I21'].value
            sheet['I'+str(cur)].font=Font(bold=True)
            sheet['J'+str(cur)]=sheet['J21'].value
            sheet['J'+str(cur)].font=Font(bold=True)
            sheet['A'+str(cur)].border = mediumBorderToBottom
            sheet['B'+str(cur)].border = mediumBorderToBottom
            sheet['C'+str(cur)].border = mediumBorderToBottom
            sheet['D'+str(cur)].border = mediumBorderToBottom
            sheet['E'+str(cur)].border = mediumBorderToBottom
            sheet['F'+str(cur)].border = mediumBorderToBottom
            sheet['G'+str(cur)].border = mediumBorderToBottom
            sheet['H'+str(cur)].border = mediumBorderToBottom
            sheet['I'+str(cur)].border = mediumBorderToBottom
            sheet['J'+str(cur)].border = mediumBorderToBottom
            tempDebitTotal=0
            tempNoOfDebit=0
            tempCreditTotal=0
            tempNoOfCredit=0
        else:pass
        transCounter=transCounter+1;
        cur=cur+2
        count=count+1;
    transCounter = transCounter-1
    if transCounter%6==0:pass
    else:
        sheet['A'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        sheet['A'+str(cur)].border = mediumBorderTop
        sheet['A'+str(cur)].font = Font(bold=True)
        sheet['A'+str(cur+1)].border = mediumBorderBottom
        sheet['G'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        sheet['G'+str(cur)].border = mediumBorderTop
        sheet['G'+str(cur)].font = Font(bold=True)
        sheet['G'+str(cur+1)].border = mediumBorderBottom
        sheet['C'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        sheet['C'+str(cur)].border = mediumBorderTop
        sheet['C'+str(cur+1)].border = mediumBorderBottom
        sheet['C'+str(cur)].font = Font(bold=True)
        sheet['I'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        sheet['I'+str(cur)].border = mediumBorderTop
        sheet['I'+str(cur)].font = Font(bold=True)
        sheet['I'+str(cur+1)].border = mediumBorderBottom
        sheet['A'+str(cur)]="No. of Debits"
        sheet['A'+str(cur+1)]=str(tempNoOfDebit)
        sheet.merge_cells("A"+str(cur)+":"+"B"+str(cur))
        sheet.merge_cells("A"+str(cur+1)+":"+"B"+str(cur+1))
        sheet['G'+str(cur)]="No. of Credits"
        sheet['G'+str(cur+1)]=str(tempNoOfCredit)
        sheet.merge_cells("G"+str(cur)+":"+"H"+str(cur))
        sheet.merge_cells("G"+str(cur+1)+":"+"H"+str(cur+1))
        sheet['C'+str(cur)]="Total Amount - Debits \n$"
        sheet['C'+str(cur+1)]=str("{:.2f}".format(float(tempDebitTotal)))
        sheet.merge_cells("C"+str(cur)+":"+"F"+str(cur))
        sheet.merge_cells("C"+str(cur+1)+":"+"F"+str(cur+1))
        sheet['I'+str(cur)]="Total Amount - Credits \n$"
        sheet['I'+str(cur+1)]=str("{:.2f}".format(tempCreditTotal))
        sheet.merge_cells("I"+str(cur)+":"+"J"+str(cur))
        sheet.merge_cells("I"+str(cur+1)+":"+"J"+str(cur+1))
        if len(result)-4>=0:
            tempCount=len(result)-4
            if tempCount%6==0:cur=cur+4
            else:
                reminder=tempCount%6
                if reminder-3==0:cur=cur+4
                elif reminder-3>0:cur=cur+1
                else:pass
        elif  len(result)-4==-1 or len(result)-4==-2 or len(result)-4==-3:
            cur = cur+3
        else:
            cur=cur+2
            #sheet['J32']=str(statementPage)+" of "+str(totalStatementPages)
        if len(result)==3:
            try:sheet['J'+str(34*int(statementPage))]=str(statementPage)+" of "+str(totalStatementPages)
            except:pass
        else:
            sheet['A'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            sheet.row_dimensions[cur].height=160
            tempText="This is your official account statement generated by us. Report any errors or omissions within 30 days of receipt electronically of this statement. Please see the terms and conditions of the applicable Scotiabank Financial Services Agreement or Business Banking Services Agreement for your account obligations.\nAll service fees and charges may be subject to any applicable sales taxes (GST/PST/QST/HST) or any tax levied by the government thereafter. These taxes will be payable by the customer.\n\nGST Registration No. R105195598\nÂ® Registered trademark of The Bank of Nova Scotia"
            sheet['A'+str(cur)]=tempText
            sheet.merge_cells("A"+str(cur)+":"+"J"+str(cur))
            cur=cur+2
            try:sheet['J'+str(34*int(statementPage))]=str(statementPage)+" of "+str(totalStatementPages)
            except:pass
        statementPage=statementPage+1
    if emptyStatement:
        sheet['A'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        sheet['A'+str(cur)].border = mediumBorderTop
        sheet['A'+str(cur)].font = Font(bold=True)
        sheet['A'+str(cur+1)].border = mediumBorderBottom
        sheet['G'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        sheet['G'+str(cur)].border = mediumBorderTop
        sheet['G'+str(cur)].font = Font(bold=True)
        sheet['G'+str(cur+1)].border = mediumBorderBottom
        sheet['C'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        sheet['C'+str(cur)].border = mediumBorderTop
        sheet['C'+str(cur+1)].border = mediumBorderBottom
        sheet['C'+str(cur)].font = Font(bold=True)
        sheet['I'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        sheet['I'+str(cur)].border = mediumBorderTop
        sheet['I'+str(cur)].font = Font(bold=True)
        sheet['I'+str(cur+1)].border = mediumBorderBottom
        sheet['A'+str(cur)]="No. of Debits"
        sheet['A'+str(cur+1)]=str(tempNoOfDebit)
        sheet.merge_cells("A"+str(cur)+":"+"B"+str(cur))
        sheet.merge_cells("A"+str(cur+1)+":"+"B"+str(cur+1))
        sheet['G'+str(cur)]="No. of Credits"
        sheet['G'+str(cur+1)]=str(tempNoOfCredit)
        sheet.merge_cells("G"+str(cur)+":"+"H"+str(cur))
        sheet.merge_cells("G"+str(cur+1)+":"+"H"+str(cur+1))
        sheet['C'+str(cur)]="Total Amount - Debits \n$"
        sheet['C'+str(cur+1)]=str("{:.2f}".format(float(tempDebitTotal)))
        sheet.merge_cells("C"+str(cur)+":"+"F"+str(cur))
        sheet.merge_cells("C"+str(cur+1)+":"+"F"+str(cur+1))
        sheet['I'+str(cur)]="Total Amount - Credits \n$"
        sheet['I'+str(cur+1)]=str("{:.2f}".format(tempCreditTotal))
        sheet.merge_cells("I"+str(cur)+":"+"J"+str(cur))
        sheet.merge_cells("I"+str(cur+1)+":"+"J"+str(cur+1))
        cur=cur+4
        sheet['A'+str(cur)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        sheet.row_dimensions[cur].height=160
        tempText="This is your official account statement generated by us. Report any errors or omissions within 30 days of receipt electronically of this statement. Please see the terms and conditions of the applicable Scotiabank Financial Services Agreement or Business Banking Services Agreement for your account obligations.\nAll service fees and charges may be subject to any applicable sales taxes (GST/PST/QST/HST) or any tax levied by the government thereafter. These taxes will be payable by the customer.\n\nGST Registration No. R105195598\nÂ® Registered trademark of The Bank of Nova Scotia"
        sheet['A'+str(cur)]=tempText
        sheet.cell(row=cur, column=5, value=tempText)
        sheet.merge_cells("A"+str(cur)+":"+"J"+str(cur))
        cur=cur+3
        sheet['J34']=str(statementPage)+" of "+str(totalStatementPages)
    try:
        sheet['A18']=noOfDebit
        sheet['G18'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        sheet['G18']=noOfCredit
        sheet['C18']="$"+str("{:.2f}".format(float(debitTotal)))
        sheet['I18']="$"+str("{:.2f}".format(creditTotal))
    except:pass
    temp=randrange(1000, 100000)
    xfile.save("/var/www/verdebooks/webapp/statements/"+str(month)+"-"+str(year)+str(temp)+".xlsx")
    url="https://verdebooks.com/statements/"+str(month)+"-"+str(year)+str(temp)+".pdf"
    try:
        convertapi.api_secret = 'XqfWIpvLatF0b12A'
        result=convertapi.convert('pdf', {'File': '/var/www/verdebooks/webapp/statements/'+str(month)+'-'+str(year)+str(temp)+'.xlsx'}, from_format = 'xlsx')
        result.file.save('/var/www/verdebooks/webapp/statements/'+str(month)+"-"+str(year)+str(temp)+'.pdf')
    except:
        url="https://verdebooks.com/statements/"+str(month)+"-"+str(year)+str(temp)+".pdf"
    return jsonify({"response":url})


@app.route("/api/transactionaAll",methods=["GET","POST"])
def transactionaAll():
    db,cursor=database()
    print("resulthjgh",)
    query="select id,date,description,withdrawal,deposite,balance from mergetransactions"
    cursor.execute(query)
    result=cursor.fetchall()
    response=[]
    print("result",result)
    if result:
        for row in result:
            try:withdraw="{:.2f}".format(row[3])
            except:withdraw=row[3]
            try:deposite="{:.2f}".format(row[4])
            except:deposite=row[4]
            try:balance="{:.2f}".format(row[5])
            except:balance=row[5]
            dateTemp=str(row[1])
            temp={"id":row[0],"date":dateTemp[0:10],"description":row[2],"withdrawal":withdraw,"deposite":deposite,"balance":balance}
            response.append(temp)
    return jsonify({"response":response})

#Current Balance

@app.route("/api/currentBalance",methods=["GET","POST"])
def currentBalance():
    db, cursor = database()
    data = request.form
    companyid = data["companyid"]
    print(companyid,"companyid")
    query="select date,balance from transactionalhistory where company_id=%s order by date desc limit 1"
    cursor.execute(query, (companyid,))
    result=cursor.fetchone()
    if result:
        query2="select date,balance from transactionalhistory where date = '"+str(result[0])+"' and company_id='"+companyid+"' order by id desc limit 1"
        cursor.execute(query2)
        result2=cursor.fetchone()
        if result2:balance="{:.2f}".format(float(result2[1]))
        else:balance="{:.2f}".format(float(result[1]))
    else:balance=0.00
    return jsonify({"balance":balance})

@app.route("/api/addBalance",methods=["GET","POST"])
def addBalance():
    data=request.form
    companyid = data["companyid"]
    db,cursor=database()
    query="select balance,id from transactionalhistory where company_id=%s order by date desc,id desc limit 1"
    cursor.execute(query,(companyid,))
    result=cursor.fetchone()
    if result:
        currentBalance=float(result[0])
    else:
        currentBalance=0
    newBalance=currentBalance+float(data["balance"])
    description="DEPOSIT HAMILTON nextON 93310 490 next47943523 MB-DEP"
    query2="insert into transactionalhistory(date,description,withdrawal,deposite,balance, company_id) values(%s,%s,%s,%s,%s,%s)"
    generateDate=str(datetime.today().strftime('%Y-%m-%d'))
    values2=(generateDate,description,"",data["balance"],newBalance, companyid)
    cursor.execute(query2,values2)
    db.commit()
    return jsonify({"response":"Added Successfully"})

#URL TO PDF
@app.route("/api/urlToPdf/<data>",methods=["GET","POST"])
def urlToPdf(data):
    api = pdfcrowd.HtmlToPdfClient("sohaib", "b68af86f1879568b27d711cd29c5346f")
    api.convertUrlToFile("https://verdebooks.com:7900/api/printStub/"+data, "/home/ubuntu/verdebooks/webapp/example1.pdf")
    return jsonify({"url":"https://verdebooks.com/example1.pdf"})

#PAY CHEQUE LIST

@app.route("/api/payChequeList", methods=["GET","POST"])
def payChequeList():
    db,cursor=database()
    query="select payDate,name,NetPay,id from stubs order by id desc"
    cursor.execute(query)
    result=cursor.fetchall()
    response=[]
    for row in result:
        if row[0]==0 or row[0]=="0":pass
        else:
            print(str(row[0]))
            try:
                try:tempDate=datetime.strptime(str(row[0]), '%m-%d-%Y').strftime('%m/%d/%y')
                except:tempDate=datetime.strptime(str(row[0]), '%Y-%m-%d').strftime('%m/%d/%y')
            except:tempDate="None"
            temp={"payDate":tempDate,"name":row[1],"totalPay":"{:.2f}".format(float(row[2])),"netPay":"{:.2f}".format(float(row[2])),"paymentMethod":"bank","id":row[3]}
            response.append(temp)
    return jsonify({"response":response})

@app.route("/api/editEmployeeProfile",methods=["GET","POST"])
def editEmployeeProfile():
    db,cursor=database()
    data=request.form
    updateParam=[]
    for k,v in enumerate(data):
        if v=="id":pass
        else:updateParam.append(v)
    query="update employee set "
    for i in updateParam:
        if i==updateParam[len(updateParam)-1]:
            query=query+str(i)+"='"+str(data[i])+"'"
        else:
            query=query+str(i)+"='"+str(data[i])+"',"
    query=query+" where id="+str(data["id"])
    cursor.execute(query)
    db.commit()
    return jsonify({"response":"We’ve saved your profile changes."})

@app.route("/api/getEmployeeRecordParam",methods=["GET","POST"])
def getEmployeeRecordParam():
    db,cursor=database()
    query="show columns from employee"
    cursor.execute(query)
    result=cursor.fetchall()
    if result:
        response=[str(row[0]) for row in result]
        return jsonify({"response":response})
    else:
        return jsonify({"response":"No data found"})

@app.route("/api/editStub",methods=["GET","POST"])
def editStub():
    db,cursor=database()
    data=request.form
    data=data.to_dict(flat=False)
    data=data["employees"][0].split(",")
    print(data)
    for i in range(0,len(data),2):
        if str(data[i+1])=='Already Generated':pass
        else:
            query="delete from stubs where id=%s"
            cursor.execute(query,(int(data[i]),))
            db.commit()
            query2="select date,withdrawal from transactionalhistory where stubId=%s"
            cursor.execute(query2,(int(data[i]),))
            result=cursor.fetchone()
            query3="delete from transactionalhistory where stubId=%s"
            cursor.execute(query3,(int(data[i]),))
            db.commit()
            print(result)
            query4="update transactionalhistory set balance=balance+%s where date > %s"
            cursor.execute(query4,(result[1],result[0]))
            db.commit()
    return jsonify({"response":data})

UPLOAD_FOLDER = 'uploads'  # Create a directory for uploads
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


description_labels = {
    "DEPOSIT": "DEPOSIT",
    "DEP" : "DEP",
    "PAYABLE": "PAYABLE",
    "MISC PAYMENT": "MISC PAYMENT",
    "PEEL AND PAYMENT": "PEEL AND PAYMENT",
    "CREDIT MEMO": "CREDIT MEMO",
}

@app.route("/api/uploadPDF", methods=['POST'])
def uploadPDF():
    db, cursor = database()
    getemploayeeid = request.form['company_id']
    file = request.files['pdfFile']
    # print(getemploayeeid, file)
    if file and file.filename.endswith('.pdf'):
        filename = secure_filename(file.filename)
        pdf_data = BytesIO(file.read())
        select_query = "SELECT name FROM bank_pdfsname WHERE name = %s"
        cursor.execute(select_query, (filename,))
        existing_filename = cursor.fetchone()
        if existing_filename:
            return jsonify({"response": "already"})
        random_number = random.randint(1, 100000)  
        pdfId = 00 + random_number
        insert_query = "INSERT INTO bank_pdfsname (name, pdfId,company_id) VALUES (%s, %s, %s)"
        cursor.executemany(insert_query, [(filename, pdfId,getemploayeeid)])
        db.commit()
        try: 
         with pdfplumber.open(pdf_data ) as pdf:
          for page in pdf.pages:
            text = page.extract_text()
            lines = text.splitlines()
            cleaned_lines = []
            i = 0
            first_entry = True
            while i < len(lines):
                line = lines[i]
                match = re.match(r'(\d{2}/\d{2}/\d{4})\s+(.*?)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})$', line)
                if match:
                    date_str = match.group(1)
                    description = match.group(2)
                    amount1 = float(match.group(3).replace(',', ''))
                    amount2 = float(match.group(4).replace(',', ''))
                    date = datetime.strptime(date_str, '%m/%d/%Y').date()
                    label = None
                    for desc, lbl in description_labels.items():
                        if desc in description.upper():
                            label = lbl
                            break
                    i += 1
                    additional_description = ""
                    while i < len(lines) and not re.match(r'\d{2}/\d{2}/\d{4}', lines[i]):
                        additional_description += " " + lines[i].strip()
                        i += 1
                    stubId = random.randint(1, 999999)
                    while True:
                        check_stub_query = "SELECT id FROM transactionalhistory WHERE stubId = %s"
                        cursor.execute(check_stub_query, (stubId,))
                        existing_stub = cursor.fetchone()
                        if not existing_stub:
                            break
                        stubId = random.randint(1, 999999)    
                    if label:
                        if description + additional_description  == "MISC PAYMENT WISE CANAD":
                            cleaned_line = (filename,date, description + additional_description, amount1,0, pdfId,stubId,getemploayeeid)
                        else:
                            cleaned_line = (filename,date, description + additional_description, 0, amount1,pdfId,stubId,getemploayeeid)
                    else:
                        cleaned_line = (filename,date, description + additional_description, amount1, 0,pdfId,stubId,getemploayeeid)
                        # print(cleaned_line)
                    cleaned_lines.append(cleaned_line)
                else:
                    i += 1
            insert_query = "INSERT INTO bank_pdfs (pdf_name, date,description, withdrawal, deposite,pdfId,stubId,company_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
            cursor.executemany(insert_query, cleaned_lines)
            db.commit()           
         db.close()        
         return jsonify({"response": "uploaded"})
        except Exception as e:
         return jsonify({'response': str(e)}), 500
    else:
        return jsonify({"response": "Invalid"})

@app.route("/api/allbankpdfs", methods=['GET', 'POST'])
def allbankpdfs():
    db, cursor = database()
    getemploayeeid = request.form['company_id']
    print(getemploayeeid,"comp")
    query = "SELECT id, name,pdfId,status FROM bank_pdfsname WHERE company_id = %s  ORDER BY id DESC"
    cursor.execute(query,(getemploayeeid,))
    result = cursor.fetchall()
    response = []
    for row in result:
        response.append({
            "id": row[0],
            "name": row[1],
            "pdfId" : row[2],
            "status" : row[3]
        })
    return jsonify({"response": response})

@app.route("/api/addtopdftransactions", methods=["GET","POST"])
def addtopdftransactions():
    if request.method == "POST":
        db, cursor = database()
        data = request.form
        company_id = data['company_id']
        pdfid = data['id']
        pdf_name = data['pdf_name'] if data['pdf_name'] else None
        date = data['date'] if data['date'] else None
        description = data['description'] if data['description'] else None
        withdrawal = data['withdrawal'] if data['withdrawal'] else 0
        deposite = data['deposite'] if data['deposite'] else 0
        stubId = random.randint(1, 999999)
        update_sql = "Insert into  bank_pdfs SET pdf_name=%s, date=%s, description=%s, withdrawal=%s, deposite=%s,pdfId=%s,stubId=%s,company_id=%s"
        cursor.execute(update_sql, (pdf_name, date, description, withdrawal, deposite,pdfid,stubId,company_id))
        db.commit()
        db.close()
        return jsonify({"response": "Data add successfully"})



@app.route("/api/eachbankpdfstransaction/<data><data1>", methods=['GET', 'POST'])
def eachbankpdfs(data, data1):
    pdfId = data
    getcompanyid = data1
    print(pdfId, getcompanyid)
    db, cursor = database()
    check_stub_query = "SELECT id FROM transactionalhistory WHERE pdfId = %s"
    cursor.execute(check_stub_query, (pdfId,))
    existing_stub = cursor.fetchone()
    if existing_stub:
        return jsonify({"response": "Transactions already saved"})
    
    # fetch pdf data
    provided_id = data
    # print(provided_id, "pro")
    query = "SELECT date, description, withdrawal, deposite, pdfId, stubId FROM bank_pdfs WHERE pdfId = %s and company_id = %s ORDER BY id ASC"
    cursor.execute(query, (provided_id,getcompanyid,))
    result1 = cursor.fetchall()
    response = []

    # fetch pdf last date
    query = "SELECT date FROM bank_pdfs WHERE pdfId = %s and company_id = %s ORDER BY id DESC LIMIT 1"
    cursor.execute(query, (provided_id,getcompanyid,))
    resultforlastpdfdate = cursor.fetchone()
    lastpdfdate = resultforlastpdfdate[0]
    print(resultforlastpdfdate, "(lastpdfdate)")
      # fetch transaction date of last month 
    query = "SELECT balance, date FROM transactionalhistory  WHERE DATE_FORMAT(date, '%Y-%m') < DATE_FORMAT(%s, '%Y-%m') and company_id = %s ORDER BY date DESC, id DESC LIMIT 1"
    cursor.execute(query, (lastpdfdate,getcompanyid,))
    lastmonthtransaction_date = cursor.fetchone()
    if lastmonthtransaction_date is not None:
      lastmonthtransaction_date_blnc = lastmonthtransaction_date[0]
    else:
        lastmonthtransaction_date_blnc = 0  
    print(lastmonthtransaction_date ,lastmonthtransaction_date_blnc, "lastmonthtransaction_date_blnc") 
    # Calculate the date for the first day of the next month
    next_month_date = datetime.strptime(str(lastpdfdate), '%Y-%m-%d')
    next_month_date = next_month_date.replace(day=1)  # Set the day to 1
    if next_month_date.month == 30:    
     next_month_date = next_month_date.replace(year=next_month_date.year + 1, month=1)
    else:    
     next_month_date = next_month_date.replace(month=next_month_date.month + 1)
    query2 = "SELECT balance,date,withdrawal,deposite FROM transactionalhistory WHERE date >= %s and company_id = %s LIMIT 1"
    cursor.execute(query2, (next_month_date,getcompanyid,))
    nextmonthtransaction_date = cursor.fetchone()
    if nextmonthtransaction_date is not None:
       nextmonthtransaction_date_blnc = nextmonthtransaction_date[0]
       nextmonthtransaction_date_withdrawal = nextmonthtransaction_date[2]
       nextmonthtransaction_date_deposite = nextmonthtransaction_date[3]
    else :    
       nextmonthtransaction_date_blnc = 0
       nextmonthtransaction_date_withdrawal = 0
       nextmonthtransaction_date_deposite = 0
    print(nextmonthtransaction_date_blnc,"nextmonthtransaction_date_blnc",nextmonthtransaction_date_withdrawal,nextmonthtransaction_date_deposite)
    # fetch update balance to zero
    update_query = "UPDATE transactionalhistory SET balance = 0 WHERE DATE_FORMAT(date, '%Y-%m') = DATE_FORMAT(%s, '%Y-%m') and company_id = %s"
    cursor.execute(update_query, (lastpdfdate,getcompanyid))
    db.commit()
    balance = lastmonthtransaction_date_blnc  
    try:        
        for row in result1:
            date = row[0]
            description = row[1]
            withdrawal = row[2]
            deposite = row[3]
            pdfId = row[4]
            stubId = row[5]
            balance = 0
            response.append({
                "date": date,
                "description": description,
                "withdrawal": withdrawal,
                "deposite": deposite,
                "pdfId": pdfId,
                "balance": balance,
                "stubId": stubId
            })
            # Execute the insert query for the current transaction
            insert_query = "INSERT INTO transactionalhistory (date, description, withdrawal, deposite, balance, pdfId, stubId, company_id) VALUES (%s, %s, %s, %s, %s, %s, %s,%s)"
            cursor.execute(insert_query, (date, description, withdrawal, deposite, balance, pdfId, stubId,getcompanyid))       
        query = "SELECT * FROM transactionalhistory WHERE balance <= %s ORDER BY date ASC"
        zeroblnc = 0
        cursor.execute(query, (zeroblnc,))
        result2 = cursor.fetchall()
        for row in result2:
         if row[4]:
           lastmonthtransaction_date_blnc += row[4]
         if row[3]:
           withdrawal = row[3]
           payDate = row[1]
           if lastmonthtransaction_date_blnc < withdrawal:
                while True:
                   rnd_balance = random.randint(5000, 80000)
                   if balance + rnd_balance >= withdrawal:
                       break
                description1 = "DEPOSIT HAMILTON nextON 93310 490 next47943523 MB-DEP"
                lastmonthtransaction_date_blnc += rnd_balance
                # print(f"Added {rnd_balance:.2f} to cover the transaction. New balance: {lastmonthtransaction_date_blnc:.2f}")
                rnd_amount_data = (payDate, description1, 0, rnd_balance, lastmonthtransaction_date_blnc, 0, 0,getcompanyid)
                insert_query = "INSERT INTO transactionalhistory (date, description, withdrawal, deposite, balance, pdfId, stubId,company_id) VALUES (%s, %s, %s, %s, %s, %s, %s,%s)"
                cursor.execute(insert_query, rnd_amount_data)
           lastmonthtransaction_date_blnc -= withdrawal
         update_query = "UPDATE transactionalhistory SET balance = %s WHERE id = %s ORDER BY date ASC"
         cursor.execute(update_query, (lastmonthtransaction_date_blnc, row[0]))
         updatestatus(getcompanyid,pdfId)
         db.commit()
         
        print("Last Data Balance:", lastmonthtransaction_date_blnc,nextmonthtransaction_date_blnc,type(lastmonthtransaction_date_blnc),type(nextmonthtransaction_date_blnc))
        
        if nextmonthtransaction_date_blnc != 0 and lastmonthtransaction_date_blnc != 0:
        # RANDOM WITHDRAW  
         if lastmonthtransaction_date_blnc > nextmonthtransaction_date_blnc:
            balance_difference = int(lastmonthtransaction_date_blnc - nextmonthtransaction_date_blnc)
            with_description  = "WITHDRAWAL HAMILTON nextON 93310 490 next47943523 MB-WITH"
            with_date = row[1]
            with_balance = lastmonthtransaction_date_blnc - balance_difference
            if nextmonthtransaction_date_withdrawal:
                 setblnc =  nextmonthtransaction_date_withdrawal + with_balance
            elif nextmonthtransaction_date_deposite:
                if nextmonthtransaction_date_deposite:
                   setblnc = with_balance - nextmonthtransaction_date_deposite  
                else:                    
                   setblnc = nextmonthtransaction_date_deposite - with_balance 
            newblnc = setblnc
            rnd_data = (with_date, with_description, balance_difference,0,  newblnc, 0, 0,getcompanyid)
            print(rnd_data,"data1")
        # RANDOM DEPOSIT
         elif lastmonthtransaction_date_blnc < nextmonthtransaction_date_blnc:
            balance_difference = int(nextmonthtransaction_date_blnc - lastmonthtransaction_date_blnc)
            dep_description= "DEPOSIT HAMILTON nextON 93310 490 next47943523 MB-DEP"
            dep_date = row[1]
            dep_balance = lastmonthtransaction_date_blnc + balance_difference
            if nextmonthtransaction_date_withdrawal:
                 setblnc =  nextmonthtransaction_date_withdrawal + dep_balance
            elif nextmonthtransaction_date_deposite:
                setblnc = nextmonthtransaction_date_deposite - dep_balance 
            newblnc = setblnc
            rnd_data = (dep_date, dep_description, 0,balance_difference, newblnc, 0, 0,getcompanyid)
            print(rnd_data,"data")
            print("Last Data Balance:", lastmonthtransaction_date_blnc,nextmonthtransaction_date_blnc,type(lastmonthtransaction_date_blnc),type(nextmonthtransaction_date_blnc))

         insert_query = "INSERT INTO transactionalhistory (date, description, withdrawal, deposite, balance, pdfId, stubId,company_id) VALUES (%s, %s, %s, %s, %s, %s, %s,%s)"
         cursor.execute(insert_query, rnd_data)        
        else:
         updatestatus(getcompanyid,pdfId)
         return jsonify({"response": "Transactions added successfully"})
        db.commit()
        return jsonify({"response": "Transactions added successfully"})       
    except Exception as e:
        return jsonify({"error": str(e)})

def updatestatus(getcompanyid,pdfId):
     db, cursor = database()
     update_query = "UPDATE bank_pdfsname SET status ='Merged' WHERE company_id = %s and pdfId = %s"
     print(update_query,"update")
     cursor.execute(update_query, (getcompanyid,pdfId,))
     db.commit()

@app.route("/api/edittransaction", methods=["GET","POST"])
def edittransaction():
    if request.method == "POST":
        db, cursor = database()
        data = request.form
        id = data['id']
        pdf_name = data['pdf_name'] if data['pdf_name'] else None
        date = data['date'] if data['date'] else None
        description = data['description'] if data['description'] else None
        withdrawal = data['withdrawal'] if data['withdrawal'] else 0
        deposite = data['deposite'] if data['deposite'] else 0
        update_sql = "UPDATE bank_pdfs SET pdf_name=%s, date=%s, description=%s, withdrawal=%s, deposite=%s WHERE id=%s"
        cursor.execute(update_sql, (pdf_name, date, description, withdrawal, deposite, id))
        db.commit()
        db.close()
        return jsonify({"response": "Data updated successfully"})

@app.route("/api/deltransaction", methods=["GET","POST"])
def deltransaction():
    if request.method == "POST":
        db, cursor = database()
        data = request.form
        id = data['id']
        print(id,"Dsf")
        sql = "DELETE FROM bank_pdfs WHERE id = %s"
        adr = (data['id'],)
        cursor.execute(sql, adr)
        db.commit()
        db.close()
        return jsonify({"response": "Data Deleted successfully"})

@app.route("/api/bankpdfs", methods=['GET', 'POST'])
def bankpdfs():
    db, cursor = database()
    data = request.form
    id = data['urlid']
    query = "SELECT id, pdf_name,date,description, withdrawal, deposite, company_id FROM bank_pdfs WHERE pdfId=%s  ORDER BY id ASC"
    cursor.execute(query,(id,))
    result = cursor.fetchall()
    response = []
    for row in result:
        response.append({
            "id": row[0],
             "pdf_name": row[1],
             "date": row[2],
             "description": row[3],
             "withdrawal": row[4],
             "deposite": row[5],
             "company_id": row[6]
        })
    return jsonify({"response": response})

@app.route("/api/stubTransactionalhistory", methods=['GET', 'POST'])
def stubTransactionalhistory():
    db, cursor = database()
    data = request.form
    company_id = data['id']
    print(company_id,"company_id")
    db, cursor = database()
    query = "SELECT DISTINCT DATE_FORMAT(payDate, '%Y-%m') AS month FROM stubs WHERE company_id = %s AND payDate != 0"
    cursor.execute(query, (company_id,))
    months = cursor.fetchall()
    unique_months = set(month[0] for month in months)
    datetime_months = [datetime.strptime(month, '%Y-%m') for month in unique_months]
    sorted_months = sorted(datetime_months, reverse=True)
    formatted_months = [month.strftime('%B %Y') for month in sorted_months]
    return jsonify({"response": formatted_months})

@app.route("/api/eachstubmonth", methods=["GET","POST"])
def eachstubmonth():
    if request.method == "POST":
        db, cursor = database()
        data = request.form
        month_str = data['month']
        company_id = data['company_id']
        print("month_str", month_str,company_id)
        try:
            month = datetime.strptime(month_str, '%B %Y')
        except ValueError:
            return "Invalid date format. Use 'Month Year' format, e.g., 'June 2023'"
        month_formatted = month.strftime('%Y-%m')
        zeropdfid = 0
        print("month_formatted", month_formatted)
        query = "SELECT id,payDate,name,NetPay FROM stubs WHERE company_id = %s AND DATE_FORMAT(payDate, '%Y-%m') = %s ORDER BY id ASC"
        cursor.execute(query, (company_id, month_formatted))
        result = cursor.fetchall()
        response = []
        for row in result:
            response.append({
                "id": row[0],
                "payDate": row[1],
                "name": row[2],
                "NetPay": row[3]
            })
        return jsonify({"response": response})



@app.route("/api/delpdf", methods=["GET","POST"])
def delpdf():
    if request.method == "POST":
        db, cursor = database()
        data = request.form
        id = data['id']
        print(id, "id")
        query = "SELECT * FROM bank_pdfsname WHERE pdfId = %s"
        cursor.execute(query, (id,))
        result = cursor.fetchall()
        print(result)
        if result:
            delete_query1 = "DELETE FROM bank_pdfs WHERE pdfID = %s"
            cursor.execute(delete_query1, (id,))
            delete_query = "DELETE FROM bank_pdfsname WHERE pdfID = %s"
            cursor.execute(delete_query, (id,))
            db.commit()
            if delete_query1:
             return jsonify({"response": "Data Deleted successfully"})
            else:
              return jsonify({"response": "Data Deleted successfully"})
        else:
            return jsonify({"response": "Record not found"})    


@app.route("/api/delstubtransaction", methods=["GET","POST"])
def delstubtransaction():
    if request.method == "POST":
        db, cursor = database()
        data = request.form
        id = data['id']
        print(id, "id")
        query = "SELECT * FROM transactionalhistory WHERE stubId = %s"
        cursor.execute(query, (id,))
        result = cursor.fetchall()
        print(result)
        if result:
            delete_query = "DELETE FROM transactionalhistory WHERE stubId = %s"
            cursor.execute(delete_query, (id,))
            delete_query = "DELETE FROM stubs WHERE id = %s"
            cursor.execute(delete_query, (id,))
            db.commit()
            return jsonify({"response": "Data Deleted successfully"})
        else:
            return jsonify({"response": "Record not found"})

# RUN APPLICATION

if __name__=="__main__":
    # db, cursor = database()
    #gevent.get_hub().SYSTEM_ERROR = BaseException
    app.secret_key="Infiniti123"
    print("server is running on 7900")
    # #context=('cert.pem','privkey.pem')
    # #app.run(port=7900, debug=True, host='0.0.0.0', ssl_context=context, threaded=True)
    # http_server = WSGIServer(('0.0.0.0', 7900), app,certfile="/etc/letsencrypt/live/verdebooks.com/fullchain.pem",keyfile="/etc/letsencrypt/live/verdebooks.com/privkey.pem")
    #context = ssl.create_default_context(ssl.Purpose.CLIENT_AUTH)
    #context.load_cert_chain(certfile='/etc/letsencrypt/live/verdebooks.com/fullchain.pem', keyfile='/etc/letsencrypt/live/verdebooks.com/privkey.pem')

    #http_server = WSGIServer(('0.0.0.0', 7900), app)
    #http_server.ssl_context = context
    #http_server.serve_forever()
 # http_server = WSGIServer(('0.0.0.0', 7900), app)
    # #http_server = WSGIServer(('', 7900),app)
    # http_server.serve_forever()
    #context = ('/etc/letsencrypt/live/verdebooks.com/fullchain.pem','/etc/letsencrypt/live/verdebooks.com/privkey.pem')
    app.run(host='0.0.0.0', port=7900)
